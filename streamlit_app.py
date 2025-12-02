import streamlit as st
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import re
import io

# ------------------------------------------------------------------
# Funções auxiliares
# ------------------------------------------------------------------
def sanitize_sheet_name(name):
    invalid_chars = r'[\/:*?[\]]'
    name = re.sub(invalid_chars, '_', str(name))
    return name[:31]

# ------------------------------------------------------------------
# Função principal de parsing
# ------------------------------------------------------------------
def parse_almoxarifado(file_content):
    consolidated_data = []
    current_almoxarifado = None
    reading_table = False

    # Regex para cabeçalho do almoxarifado
    almoxarifado_pattern = re.compile(
        r'Almoxarifado:§*(\d+)\s*-\s*([^-§]+)\s*-\s*([^-§]+)\s*-\s*([^-§]+)\s*-\s*([^-§]+)§+'
    )

    # Regex
    data_pattern = re.compile(
        r'(\d+)§+'                     # grupo 1 → Item 
        r'(\d+)\s*-\s*([^§]+)§+'       # grupo 2 → Código SIGBP, grupo 3 → Material
        r'(\w+)§+'                     # grupo 4 → Unidade de Medida
        r'([^§]+)§+'                   # grupo 5 → Finalidade da Compra
        r'(\w+)§+'                     # grupo 6 → Endereço
        r'([\d,.]+)§+'                 # grupo 7 → Qtd Indisponível
        r'([\d,.]+)§+'                 # grupo 8 → Valor Indisponível
        r'([\d,.]+)§+'                 # grupo 9 → Qtd Disponível
        r'([\d,.]+)§+'                 # grupo 10 → Valor Disponível
        r'([\d,.]+)§+'                 # grupo 11 → Qtd Total
        r'([\d,.]+)§+'                 # grupo 12 → Valor Total
    )

    lines = file_content.decode('utf-8').splitlines()
    for raw_line in lines:
        line = raw_line.strip()
        if not line:
            reading_table = False
            continue

        # ---------- Cabeçalho do Almoxarifado ----------
        almox_match = almoxarifado_pattern.match(line)
        if almox_match:
            #Número do almoxarifado (grupo 1) e o nome completo (grupo 5) :)
            num_almox = almox_match.group(1).strip()
            nome_almox = almox_match.group(5).strip()
            current_almoxarifado = f"{num_almox} - {nome_almox}"
            reading_table = True
            continue

        # ---------- Linhas de itens ----------
        if reading_table:
            data_match = data_pattern.match(line)
            if data_match:
                (
                    item,                  # 1
                    cod_sigbp,             # 2
                    material,              # 3
                    unidade_medida,        # 4
                    finalidade_compra,     # 5
                    endereco,              # 6
                    qtd_indisponivel,      # 7
                    valor_indisponivel,    # 8
                    qtd_disponivel,        # 9
                    valor_disponivel,      # 10
                    qtd_total,             # 11
                    valor_total            # 12
                ) = data_match.groups()

                consolidated_data.append({
                    "Almoxarifado": current_almoxarifado,
                    "Item": item.strip(),                     # ← nova coluna, depois da alteração
                    "Código": cod_sigbp.strip(),
                    "Material": material.strip(),
                    "U.M.": unidade_medida.strip(),
                    "Qtd total": qtd_total.strip().replace('.', '').replace(',', '.'),
                    "Valor total": valor_total.strip().replace('.', '').replace(',', '.'),
                    "valor unitário": None,
                    "Incorporação/Baixa": None,
                    "Quantidade e Levantamento": 0.0
                })

    # ----------------------- DataFrame -----------------------
    df = pd.DataFrame(consolidated_data)

    # Conversão numérica (Só pra garantir)
    for col in ["Qtd total", "Valor total"]:
        df[col] = pd.to_numeric(df[col], errors='coerce')

    # Valor unitário
    df["valor unitário"] = df["Valor total"] / df["Qtd total"]

    # ----------------------- Excel -----------------------
    wb = Workbook()
    wb.remove(wb.active)                     # remove aba padrão

    # Estilos
    header_fill = PatternFill(start_color="006400", end_color="006400", fill_type="solid")
    green_fill  = PatternFill(start_color="DAF2D0", end_color="DAF2D0", fill_type="solid")
    white_fill  = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    gray_fill   = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    # Agrupar por almoxarifado e criar uma aba para cada um
    for almox_nome, grupo in df.groupby("Almoxarifado"):
        sheet_name = sanitize_sheet_name(almox_nome)
        ws = wb.create_sheet(title=sheet_name)

        # Escrever cabeçalho + dados
        for r_idx, row in enumerate(dataframe_to_rows(grupo, index=False, header=True), 1):
            ws.append(row)

            # Estilizar cabeçalho (primeira linha)
            if r_idx == 1:
                for cell in ws[r_idx]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = center_align
                    cell.border = thin_border
            else:
                # Cores alternadas nas colunas A-H (1-8)
                for c in range(1, 9):
                    cell = ws.cell(row=r_idx, column=c)
                    cell.fill = green_fill if r_idx % 2 == 0 else white_fill
                    cell.alignment = center_align
                    cell.border = thin_border

                # Colunas I e J (Incorporação/Baixa e Quantidade e Levantamento) → cinza
                for c in range(9, 11):
                    cell = ws.cell(row=r_idx, column=c)
                    cell.fill = gray_fill
                    cell.alignment = center_align
                    cell.border = thin_border

        # Fórmula na coluna Incorporação/Baixa (coluna I, índice 9)
        col_qtd_total = 6          # coluna F → Qtd total
        col_inc_baixa = 9          # coluna I
        for row in range(2, len(grupo) + 2):
            ws.cell(row=row, column=col_inc_baixa).value = f"=J{row}-F{row}"
            ws.cell(row=row, column=col_inc_baixa).border = thin_border

        # Ajuste automático de largura das colunas
        for col in ws.columns:
            max_length = 0
            column_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

    # Salvar em buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return df, buffer

# ------------------------------------------------------------------
# Interface Streamlit
# ------------------------------------------------------------------
st.title("Processador de Estoque Analítico")
st.write("Faça upload do arquivo .txt para gerar a tabela consolidada em Excel (com abas por almoxarifado e coluna **Item**).")

uploaded_file = st.file_uploader("Escolha um arquivo .txt", type="txt")

if uploaded_file is not None:
    try:
        df_consolidado, excel_buffer = parse_almoxarifado(uploaded_file.read())

        st.write("### Pré-visualização da tabela consolidada")
        st.dataframe(df_consolidado)

        st.download_button(
            label="Baixar Excel (com abas por almoxarifado)",
            data=excel_buffer,
            file_name="estoque_consolidado_com_item.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        st.success("Arquivo gerado com sucesso! Cada almoxarifado está em uma aba separada e a coluna **Item** foi adicionada.")
    except Exception as e:
        st.error(f"Erro ao processar o arquivo: {e}")
